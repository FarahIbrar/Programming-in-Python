#!/usr/bin/env python
# coding: utf-8

# <div style="background-color:beige;color:beige">
# <header>
# <h1 style="padding:1em;text-align:center;color:#00008B">Basics of Python programming <br><br> &nbsp;&nbsp; File Handling </h1> 
# </header>
# <br><br><br><br><br><br>
# <footer>
# <table style="border:none;width:100%">
# <tr><td style= "text-align:left;background-color:blue;color:white;font-size:80%;"> Farah Ibrar </td><td style= "text-align:right;background-color:blue;color:white;font-size:80%;"> Programming in Python </td></tr></table>
# </footer>
# 
# </div>

# <div style="background-color:beige;color:beige">
# <header>
# <h1 style="padding:1em;text-align:left;color:#00008B">Outline</h1>
# </header>
# 
#     
# <ul><li><span style="color:#00008B;font-size:24px">Open and Close a File </span> <br><br>
#         <li><span style="color:#00008B; font-size:24px">Read a File </span> <br><br>
#         <li><span style="color:#00008B; font-size:24px">Write in a File </span><br><br>
#         <li><span style="color:#00008B; font-size:24px">Append in a File </span><br><br>
#         <li><span style="color:#00008B; font-size:24px">Create a new File </span><br><br>
#         </li>
#        </ul>
# <br><br>
# <footer>
# <table style="border:none;width:100%">
# <tr><td style= "text-align:left;background-color:blue;color:white;font-size:80%;"> Farah Ibrar </td><td style= "text-align:right;background-color:blue;color:white;font-size:80%;"> Programming in Python </td></tr></table>
# </footer>
# 
# </div>

# <div style="background-color:beige;color:beige">
# 
# <h2 style="padding:1em;text-align:left;color:#00008B">Open a File</h2>
# <ul><li><span style="color:#00008B;font-size:20px">
# <b>open( )</b> - to open a file. Need to define the following two parameters - filename, mode.  <br><br>
# <b>close( )</b> - to close a file. <br><br>
# <li><span style="color:#00008B;font-size:16px">
# <i>Modes for opening a file: </i><br>
#  1. "r" (Read is the default): Opens a file for reading. Will give an error if the file does not exist. <br><br>
#  2. "w" (Write): Opens a file for writing. Creates the file if it does not exist. Will overwrite any existing content <br><br>
#  3. "a" (Append): Opens a file for appending. Creates the file if it does not exist. Will append at the end of the existing content. <br><br>
#  4. "x" (Create): Creates the specified file. Returns an error if the file exists. <br><br>
# </span></li>
# </span></li></ul>
#  
# <br>
# <footer>
# <table style="border:none;width:100%">
# <tr><td style= "text-align:left;background-color:blue;color:white;font-size:80%;"> Farah Ibrar </td><td style= "text-align:right;background-color:blue;color:white;font-size:80%;"> Programming in Python </td></tr></table>
# </footer>

# In[1]:


# Define the path to your desktop
desktop_path = "/Users/farah/Desktop/"

# Define the file name
filename = "test.txt"

# Create the full file path
full_path = desktop_path + filename

# Open the file in write mode
file1 = open(full_path, "w")

# Write some content to the file
file1.write("This is a test file created on the Mac desktop.")

# Close the file
file1.close()

# Output the file object to confirm creation
file1


# 1. Define Desktop Path: Set the path to your desktop.
# 2. Define Filename: Specify the name of the file.
# 3. Create Full Path: Combine the desktop path and filename.
# 4. Open File: Open the file in write mode ("w").
# 5. Write to File: Write a test message to the file.
# 6. Close File: Close the file to save changes.
# 7. Output: Return the file object to confirm creation.

# In[6]:


# Define the path to your desktop
desktop_path = "/Users/farah/Desktop/"

# Define the file name
filename = "test.txt"

# Create the full file path
full_path = desktop_path + filename

# Open the file in read mode
file1 = open(full_path, "r")

# Read the content of the file
content = file1.read()

# Close the file
file1.close()

# Print the content to confirm
print(content)


# In[9]:


# Open a file - by default is handled as a text and opens in a read mode (for binary mode - specify mode as "b")


# In[8]:


content = open("/Users/farah/Desktop/test.txt", "r").read()
print(open("/Users/farah/Desktop/test.txt", "r").read())

# Close a file
file1.close()


# In[ ]:


# Open a excel file
# One can open a file without any library installation, however to make it from scratch a library would be needed.  


# - You need to install and import `openpyxl` (or another library like `xlrd`) because Python's standard library does not have `built-in support` for `reading` or `writing` `Excel files`.

# In[10]:


import openpyxl

# Define the path to your desktop
desktop_path = "/Users/farah/Desktop/"

# Define the file name
filename = "test.xlsx"

# Create the full file path
full_path = desktop_path + filename

# Create a new workbook and add a worksheet
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Sheet1"

# Write some content to the worksheet
sheet["A1"] = "Hello"
sheet["B1"] = "World"

# Save the workbook
workbook.save(full_path)

# Open the workbook
workbook = openpyxl.load_workbook(full_path)
sheet = workbook.active

# Read the content from the worksheet
content = []
for row in sheet.iter_rows(values_only=True):
    content.append(row)

# Print the content
print(content)


# <div style="background-color:beige;color:beige">
# 
# <h2 style="padding:1em;text-align:left;color:#00008B">Read a File</h2>
# <ul><li><span style="color:#00008B;font-size:20px">
# "r" (Read is the default): Opens a file for reading. Will give an error if the file does not exist. <br><br>
# <b>read()</b> - by default reads all content of the file. Can specify number of characters to read. <br><br>
# <b>readline</b>	- reads line of a file. <br><br>
# <b>read-lines</b> - returns a list of lines.<br><br>
# </span></li></ul>
#  
# <br>
# <footer>
# <table style="border:none;width:100%">
# <tr><td style= "text-align:left;background-color:blue;color:white;font-size:80%;"> Farah Ibrar </td><td style= "text-align:right;background-color:blue;color:white;font-size:80%;"> Programming in Python </td></tr></table>
# </footer>

# In[11]:


# Read a file
file1 = open("/Users/farah/Desktop/test.txt")

print(file1.read)


# In[12]:


# Read a file, readline()

file1 = open("/Users/farah/Desktop/test.txt")
print(file1.readline())
# Reads the first line of the opened file.


# In[13]:


# Read a file, readlines()

file1 = open("/Users/farah/Desktop/test.txt")
file1.readlines()
# Reads all lines in the opened file.


# In[14]:


file1 = open("/Users/farah/Desktop/test.txt")

# Read only a part of file - specify how many characters
print(file1.read(4))


# In[19]:


file1.close()


# <div style="background-color:beige;color:beige">
# 
# <h2 style="padding:1em;text-align:left;color:#00008B">Write in a File</h2>
# <ul><li><span style="color:#00008B;font-size:20px">
# "w" (Write): Opens a file for writing. Creates the file if it does not exist. Will overwrite any existing content <br><br>
# </span></li></ul>
#  
# <br>
# <footer>
# <table style="border:none;width:100%">
# <tr><td style= "text-align:left;background-color:blue;color:white;font-size:80%;"> Farah Ibrar </td><td style= "text-align:right;background-color:blue;color:white;font-size:80%;"> Programming in Python </td></tr></table>
# </footer>

# In[22]:


# Open a file - write
file1 = open("/Users/farah/Desktop/test.txt", "w")

file1


# In[25]:


# Open the file in write mode
file1 = open("/Users/farah/Desktop/test.txt", "w")

# Write the first line
file1.write("My name is Farah.\n")

# \n is for new line
# Write the second line
file1.write("I live in New York City.\n")

# Write the third line
file1.write("I love playing tennis.\n")

# Close the file
file1.close()

# Open the file again in read mode
file1 = open("/Users/farah/Desktop/test.txt", "r")

# Read and print all lines
content = file1.read()
print(content)

# Close the file
file1.close()


# In[26]:


file1 = open ("/Users/farah/Desktop/test.txt", "r")
file1


# In[28]:


file1.readline()


# In[33]:


file1.close()
file1 = open ("/Users/farah/Desktop/test.txt", "w")
file1.write("I love london. \n")


# In[35]:


file1 = open ("/Users/farah/Desktop/test.txt", "r")
file1.readlines()


# In[37]:


# Define the content to write to the file
content_to_write = [
    'My name is Farah.',
    'I live in New York City.',
    'I love playing tennis.',
    'I love cricket as well.'
]

# Open the file in write mode
file1 = open("/Users/farah/Desktop/test.txt", "w")

# Write each line from the list to the file
for line in content_to_write:
    file1.write(line + "\n")

# Close the file
file1.close()

# Open the file again in read mode
file1 = open("/Users/farah/Desktop/test.txt", "r")

# Read and print all lines
content = file1.readlines()
for line in content:
    print(line.strip())  # strip() is used to remove the newline character

# Close the file
file1.close()


# <div style="background-color:beige;color:beige">
# 
# <h2 style="padding:1em;text-align:left;color:#00008B">Append in a File</h2>
# <ul><li><span style="color:#00008B;font-size:20px">
# "a" (Append): Opens a file for appending. Creates the file if it does not exist. Will append at the end of the existing content. <br><br>
# </span></li></ul>
#  
# <br>
# <footer>
# <table style="border:none;width:100%">
# <tr><td style= "text-align:left;background-color:blue;color:white;font-size:80%;"> Farah Ibrar </td><td style= "text-align:right;background-color:blue;color:white;font-size:80%;"> Programming in Python </td></tr></table>
# </footer>

# In[39]:


# Open a file - append
file1 = open ("/Users/farah/Desktop/test.txt", "a")
file1


# In[40]:


# Append to the file
file1.write("I am learning Python. \n")

file1.write("This is file handling. \n")

# Close the file
file1.close()


# In[41]:


file1 = open ("/Users/farah/Desktop/test.txt", "r")
file1.readlines()


# In[43]:


# Define the list to append
append_list = [8, 16, 6, 1, 7]

# Convert the list to a string
append_str = ' '.join(map(str, append_list)) + '\n'

# Open the file in append mode ('a')
file1 = open("/Users/farah/Desktop/test.txt", "a")

# Append the list to the file
file1.write(append_str)

# Close the file
file1.close()

# Open the file again in read mode
file1 = open("/Users/farah/Desktop/test.txt", "r")

# Read and print all lines
content = file1.readlines()
for line in content:
    print(line.strip())  # strip() is used to remove the newline character

# Close the file
file1.close()


# `append_str = ' '.join(map(str, append_list)) + '\n'`
# Converts the list append_list to a string format with numbers separated by spaces and adds a newline character (\n) at the end.
# 
# `file1.write(append_str)`
# Appends the string representation of append_list to the file.

# In[46]:


# Original list
list = [22, 56, 54, 47, 51, 60]

# Open file1 in append mode
with open('file1.txt', 'a') as file1:
    # Write the new list to file1
    for x in list:
        file1.write("%d\n" % x)

# file1 is automatically closed at the end of the with statement


# In[47]:


# Open file1 in read mode
with open('file1.txt', 'r') as file1:
    # Read all lines from the file
    lines = file1.readlines()

# Print all lines
for line in lines:
    print(line.strip())  # strip() removes any extra newline characters

# file1 is automatically closed at the end of the with statement


# In[49]:


# Open file1 in read mode
with open('file1.txt', 'r') as file1:
    # Read the entire content of the file
    content = file1.read()

# Print the entire content
print(content)

# file1 is automatically closed at the end of the with statement


# In[50]:


import os

# Get the absolute path of file1.txt
file_path = os.path.abspath('file1.txt')
print("/Users/farah/Desktop/test.txt:", file_path)


# - File path got mixed up so fixed it using `os`.

# In[52]:


import os

# Original list
my_list = [1, 2, 3, 4, 5, 6]

# Define the file path
file_path = '/Users/farah/Desktop/test.txt'

# Open the file in append mode
with open(file_path, 'a') as file1:
    # Write the new list to the file
    for x in my_list:
        file1.write("%d\n" % x)

# Read the entire content of the file and print it
with open(file_path, 'r') as file1:
    content = file1.read()
    print("Content of test.txt:")
    print(content)

# file1 is automatically closed at the end of the with statement


# In[ ]:


# 
get_ipython().run_line_magic('s=', 'string (values will be written as string)')
get_ipython().run_line_magic('d=', 'numbers (vlaues will be written as numbers)')

if you have mixture of string and numbers then it can be run as string. 

make sure to close the file so you can see the outcome. 


# In[53]:


file1 = open ('/Users/farah/Desktop/test.txt', 'a')
file1


# In[54]:


# Open the file in append mode
with open('/Users/farah/Desktop/test.txt', 'a') as file1:
    # Define the new list
    newlist = [55, 43, 65, 88, 42]

    # Square each number and write to the file
    for x in newlist:
        squared = x ** 2
        file1.write("%d\n" % squared)

# Read the entire content of the file and print it
with open('/Users/farah/Desktop/test.txt', 'r') as file1:
    content = file1.read()
    print("Content of test.txt:")
    print(content)

# file1 is automatically closed at the end of the with statement


# 1. **Opening the File in Append Mode**:
#    ```python
#    with open('/Users/farah/Desktop/test.txt', 'a') as file1:
#    ```
#    - This line opens the file `test.txt` located at `/Users/farah/Desktop/` in append mode (`'a'`). The `with` statement ensures that the file is properly closed after all operations are completed.
# 
# 2. **Defining the New List**:
#    ```python
#    newlist = [55, 43, 65, 88, 42]
#    ```
#    - This line defines a new list `newlist` containing integers.
# 
# 3. **Writing Squared Numbers to the File**:
#    ```python
#    for x in newlist:
#        squared = x ** 2
#        file1.write("%d\n" % squared)
#    ```
#    - This loop iterates through each number `x` in `newlist`.
#    - It calculates the square of `x` using `x ** 2`.
#    - It writes the squared value to `file1` using `file1.write("%d\n" % squared)`. `%d` is used to format the integer value of `squared`, and `"\n"` ensures each squared number is written on a new line.
# 
# 4. **Reading and Printing the Content of the File**:
#    ```python
#    with open('/Users/farah/Desktop/test.txt', 'r') as file1:
#        content = file1.read()
#        print("Content of test.txt:")
#        print(content)
#    ```
#    - This code block opens `test.txt` again, this time in read mode (`'r'`).
#    - It reads the entire content of the file using `file1.read()` and stores it in the variable `content`.
#    - It prints the content of `test.txt` to the console.
# 
# 5. **Automatic File Closure**:
#    - Both instances of opening the file (`file1`) are within `with` statements. This ensures that the file is automatically closed at the end of each `with` block, preventing any potential issues with file locks or leaks.

# <div style="background-color:beige;color:beige">
# 
# <h2 style="padding:1em;text-align:left;color:#00008B">Create a new File</h2>
# <ul><li><span style="color:#00008B;font-size:20px">
# "x" (Create): Creates the specified file. Returns an error if the file exists. <br><br>
# </span></li></ul>
#  
# <br>
# <footer>
# <table style="border:none;width:100%">
# <tr><td style= "text-align:left;background-color:blue;color:white;font-size:80%;"> Farah Ibrar </td><td style= "text-align:right;background-color:blue;color:white;font-size:80%;"> Programming in Python </td></tr></table>
# </footer>

# In[58]:


# Create a new file
file2 = open('/Users/farah/Desktop/test_fabz.txt', "x")

file2.close()
# This file can then be used to do anything...


# <div style="background-color:beige;color:beige">
# 
# <h2 style="padding:1em;text-align:left;color:#00008B">Summary</h2>
# 
# <ul><li><span style="color:#00008B;font-size:20px">
# File handling in Python - open, close, read, write, append, create. <br><br>
# A file by default is handled in text mode, for binary mode use the mode "b" (e.g. - for images).<br><br>
# <a href="https://docs.python.org/3/tutorial/inputoutput.html#reading-and-writing-files">Reading and Writing Files - Resource 1</a> <br><br>
# <a href="https://realpython.com/read-write-files-python/">Reading and Writing Files - Resource 2</a><br><br>
# <a href="https://docs.python.org/3/library/csv.html">CSV File Reading and Writing - Resource 3</a><br><br>
# </span></li>
#     </ul>
# 
# <br><br><br><br>
# <footer>
# <table style="border:none;width:100%">
# <tr><td style= "text-align:left;background-color:blue;color:white;font-size:80%;"> Farah Ibrar</td><td style= "text-align:center;background-color:blue;color:white;font-size:80%;"> Programming in Python </td><td style= "text-align:right;background-color:blue;color:white;font-size:80%;">SCM7047</td></tr></table>
# </footer>
# 
# </div>

# <div style="background-color:beige;color:beige">
# 
# <h1 style="padding:1em;text-align:center;color:#00008B">EXERCISE</h1>
# 
# <ul><li><span style="color:#00008B;font-size:20px">
# 1. Write python code(s) to perform the following tasks:<br><br>
#   <li><span style="color:#00008B;font-size:16px">
#    a. Create a file named "my_first_file.txt". <br><br>
#    b. Write the following lines in this file [Line 1 - "The first argument is a string containing the filename." Line 2 - "The second argument is another string containing a few characters describing the way in which the file will be used."] <br><br>
#    c. Print the first five characters of this file. <br><br>
#    d. Append the following line in this file [Line - "In text mode, the default when reading is to convert platform-specific line endings (\n on Unix, \r\n on Windows) to just \n."]<br><br>
#    e. Print all the lines of this file.
#   </span></li>
#     
# </span></li></ul>
#     
# <footer>
# <table style="border:none;width:100%">
# <tr><td style= "text-align:left;background-color:blue;color:white;font-size:80%;"> Farah Ibrar </td><td style= "text-align:right;background-color:blue;color:white;font-size:80%;"> Programming in Python </td></tr></table>
# </footer>
# 
# </div>

# In[59]:


import os

# Define the file path on your desktop
file_path = "/Users/farah/Desktop/my_first_file.txt"

# Task a: Create a file named "my_first_file.txt"
with open(file_path, 'w') as file:
    file.write("The first argument is a string containing the filename.\n")
    file.write("The second argument is another string containing a few characters describing the way in which the file will be used.\n")

# Task b: Print the first five characters of this file
with open(file_path, 'r') as file:
    first_five = file.read(5)
    print(f"First five characters of the file: {first_five}")

# Task d: Append the following line in this file
with open(file_path, 'a') as file:
    file.write("In text mode, the default when reading is to convert platform-specific line endings (\\n on Unix, \\r\\n on Windows) to just \\n.\n")

# Task e: Print all the lines of this file
with open(file_path, 'r') as file:
    all_lines = file.readlines()
    print("\nAll lines of the file:")
    for line in all_lines:
        print(line.strip())  # strip() removes any extra newline characters

# file is automatically closed at the end of the with statement

